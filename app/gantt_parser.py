import csv
import io
import re
from datetime import datetime, date, timedelta
import openpyxl

AVATAR_COLORS = [
    "#3B82F6", "#8B5CF6", "#EC4899", "#10B981",
    "#F59E0B", "#06B6D4", "#6366F1", "#14B8A6", "#F97316"
]

MONTHS_MAP = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "may": 5, "jun": 6, "june": 6,
    "jul": 7, "july": 7, "aug": 8, "august": 8, "sep": 9, "september": 9, "sept": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12
}

HEADER_MAPPINGS = {
    "task_no": [
        "task no", "task_no", "s.no", "sno", "s no", "sl no", "sl.no", "task id", "item no", "item_no",
        "no", "#", "id", "wbs", "item #", "activity #", "line #", "pos", "position", "sr.no", "sr no"
    ],
    "title": [
        "activity / process step", "activity name", "process step", "process", "activity",
        "task name", "task_name", "tasks", "task", "action item", "action items", "action", "actions",
        "work item", "work items", "work", "item name", "items", "item",
        "title", "name", "deliverable", "deliverables", "step name", "steps", "step",
        "scope", "scope of work", "work package", "feature", "features", "todo", "to-do", "to do",
        "topic", "job", "milestone", "plan", "story", "user story", "requirement", "requirements",
        "content", "objective", "objectives", "description", "details", "task description", "activity description"
    ],
    "assignee": [
        "task owner", "owner", "assignee", "assigned to", "assigned_to", "resource",
        "resource name", "resources", "member", "members", "person", "who", "lead", "assigned",
        "responsible person", "responsible party", "pic", "p.i.c", "p.i.c.", "poc", "point of contact",
        "executor", "developer", "engineer", "action by", "assigned by", "in charge"
    ],
    "department": [
        "responsible", "department", "dept", "team", "function", "group", "division", "discipline"
    ],
    "start_date": [
        "start", "start date", "start_date", "begin", "begin date",
        "from", "start time", "startdate", "planned start", "commence", "commence date",
        "target start", "baseline start", "orig start", "start timeline", "start_dt", "start_time"
    ],
    "due_date": [
        "end", "end date", "end_date", "due date", "due_date", "finish",
        "finish date", "to", "deadline", "target date", "duedate", "enddate",
        "planned finish", "target finish", "target completion", "completion date",
        "target", "baseline finish", "orig finish", "end timeline", "due_dt", "end_time"
    ],
    "status": [
        "status", "state", "progress", "stage", "% complete", "percent complete",
        "% done", "progress %", "completion", "completion %", "task status", "condition", "overall status"
    ],
    "priority": [
        "priority", "severity", "urgency", "importance", "level", "rank", "prio"
    ],
    "estimated_hours": [
        "estimated hours", "est hours", "est_hours", "estimated time",
        "duration (hours)", "duration (days)", "duration", "hours", "work", "days",
        "work (hours)", "effort", "time (hrs)", "est duration", "est time", "mandays", "man-days", "manhours", "man-hours"
    ],
    "sprint": [
        "sprint", "phase", "milestone", "category", "wbs", "epic", "stream", "stage name", "section", "module", "workstream"
    ],
    "description": [
        "remark", "remarks", "description", "notes", "details", "comments", "summary", "scope notes", "observation", "observations"
    ],
    "tags": [
        "tags", "tag", "labels", "keywords", "category", "type"
    ]
}

def normalize_header(header_str):
    if not header_str:
        return ""
    return re.sub(r"[^a-z0-9% /_#\.]+", " ", str(header_str).lower()).strip()

def match_column_name(clean_header):
    if not clean_header:
        return None
    
    clean_str = clean_header.strip()

    # If the text ends with a number (e.g. "Step 1", "Reaction Step 2"), it is task data, not a column header
    if re.search(r"\d+$", clean_str) and not any(clean_str == p for p in ["p0", "p1", "p2", "p3", "p4", "wbs", "task no", "sl no"]):
        return None

    # 1. Exact match across all aliases (longest alias first)
    all_exact = []
    for canonical_field, aliases in HEADER_MAPPINGS.items():
        for alias in aliases:
            all_exact.append((alias, canonical_field))
    
    all_exact.sort(key=lambda x: len(x[0]), reverse=True)
    for alias, field in all_exact:
        if clean_str == alias:
            return field

    # 2. Word-boundary sub-phrase match with strict field prioritization
    # Only match sub-phrases if clean header is concise (<= 4 words) to avoid matching full task titles
    if len(clean_str.split()) <= 4:
        field_order = ["start_date", "due_date", "status", "priority", "estimated_hours", "assignee", "sprint", "department", "description", "tags", "task_no", "title"]
        for field in field_order:
            aliases = sorted(HEADER_MAPPINGS[field], key=len, reverse=True)
            for alias in aliases:
                if re.search(r"\b" + re.escape(alias) + r"\b", clean_str):
                    return field

    return None

def parse_date_value(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    
    # Check numeric Excel serial dates (e.g. 45520 -> 2024-08-16)
    if isinstance(val, (int, float)):
        if 20000 <= val <= 80000:
            try:
                base_date = datetime(1899, 12, 30)
                res_date = base_date + timedelta(days=float(val))
                return res_date.strftime("%Y-%m-%d")
            except Exception:
                pass
        return None

    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "n/a", "na", "-", "--", "tbd", "undeclared", "not declared", "nil", "pending"):
        return None

    # If it is a string of numeric serial date like '45520' or '45520.0'
    if re.match(r"^\d{5}(\.\d+)?$", val_str):
        try:
            num = float(val_str)
            if 20000 <= num <= 80000:
                base_date = datetime(1899, 12, 30)
                return (base_date + timedelta(days=num)).strftime("%Y-%m-%d")
        except Exception:
            pass

    # Standardize separators
    cleaned = val_str.replace(".", "-").replace("/", "-")

    # 1. YYYY-MM-DD or YYYY-M-D
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", cleaned)
    if m:
        y, mo, d = m.groups()
        try:
            return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
        except Exception:
            pass

    # 2. DD-MM-YYYY or D-M-YYYY
    m = re.search(r"(\d{1,2})-(\d{1,2})-(\d{4})", cleaned)
    if m:
        d, mo, y = m.groups()
        d_i, mo_i = int(d), int(mo)
        if mo_i > 12 and d_i <= 12:
            d_i, mo_i = mo_i, d_i
        try:
            return f"{int(y):04d}-{mo_i:02d}-{d_i:02d}"
        except Exception:
            pass

    # 3. DD-MM-YY or D-M-YY (2-digit year)
    m = re.search(r"(\d{1,2})-(\d{1,2})-(\d{2})(?!\d)", cleaned)
    if m:
        d, mo, y = m.groups()
        y_full = 2000 + int(y) if int(y) < 70 else 1900 + int(y)
        d_i, mo_i = int(d), int(mo)
        if mo_i > 12 and d_i <= 12:
            d_i, mo_i = mo_i, d_i
        try:
            return f"{y_full:04d}-{mo_i:02d}-{d_i:02d}"
        except Exception:
            pass

    # 4. Text month: '15-Aug-2026', '15 Aug 2026', 'Aug 15 2026', '15-August-2026', '15-Aug-26'
    text_cleaned = re.sub(r"[,]+", " ", val_str)
    tokens = re.split(r"[\s\-]+", text_cleaned)
    if len(tokens) >= 3:
        mo_num = None
        d_num = None
        y_num = None
        for t in tokens:
            t_low = t.lower()
            if t_low in MONTHS_MAP:
                mo_num = MONTHS_MAP[t_low]
            elif re.match(r"^\d{4}$", t):
                y_num = int(t)
            elif re.match(r"^\d{1,2}$", t):
                if d_num is None:
                    d_num = int(t)
                elif y_num is None:
                    y_num = 2000 + int(t) if int(t) < 70 else 1900 + int(t)
        if mo_num and d_num and y_num:
            try:
                return f"{y_num:04d}-{mo_num:02d}-{d_num:02d}"
            except Exception:
                pass

    return None

def normalize_status(val):
    if val is None:
        return "todo"
    
    if isinstance(val, (int, float)):
        if val >= 1.0 or val == 100:
            return "done"
        elif val > 0:
            return "in_progress"
        return "todo"

    s = str(val).lower().strip()
    if s.endswith("%"):
        try:
            num = float(s[:-1])
            if num >= 100: return "done"
            elif num > 0: return "in_progress"
            return "todo"
        except Exception:
            pass

    if any(k in s for k in ["done", "complete", "finish", "closed", "pass", "resolved", "delivered"]):
        return "done"
    elif any(k in s for k in ["delay", "in progress", "progress", "active", "working", "ongoing", "started", "doing", "wip"]):
        return "in_progress"
    elif any(k in s for k in ["review", "testing", "qa", "verify", "audit", "hold", "pending review"]):
        return "in_review"
    elif any(k in s for k in ["backlog", "future", "icebox", "wishlist", "planned"]):
        return "backlog"
    return "todo"

def normalize_priority(val, status="todo", remark=""):
    if status == "in_progress" and "delay" in str(val or "").lower():
        return "urgent"
    if remark and ("risk" in remark.lower() or "critical" in remark.lower() or "breakdown" in remark.lower()):
        return "urgent"

    if not val:
        return "medium"
    s = str(val).lower().strip()
    if any(k in s for k in ["urgent", "critical", "blocker", "p0", "p1", "highest", "immediate", "delay"]):
        return "urgent"
    elif any(k in s for k in ["high", "major", "p2", "important"]):
        return "high"
    elif any(k in s for k in ["low", "minor", "trivial", "p4", "lowest"]):
        return "low"
    return "medium"

def normalize_estimated_hours(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    
    s = str(val).lower().strip()
    if not s:
        return 0.0

    days_match = re.search(r"([\d\.]+)\s*(d|day|days|manday|mandays)", s)
    if days_match:
        try:
            return round(float(days_match.group(1)) * 8.0, 1)
        except Exception:
            pass

    hrs_match = re.search(r"([\d\.]+)", s)
    if hrs_match:
        try:
            return round(float(hrs_match.group(1)), 1)
        except Exception:
            pass

    return 0.0

def parse_gantt_file(file_bytes, filename=""):
    """
    Universal Spreadsheet & Gantt Chart Parser for ProjectPulse.
    Accepts ANY Excel (.xlsx, .xlsm, .xls) or CSV / TSV file.
    Supports multi-sheet scanning, section hierarchies, custom date styles, and undeclared timeline preservation.
    """
    is_excel = filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xls")) or file_bytes[:4] == b"PK\x03\x04"
    
    if is_excel:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        best_sheet_parsed = None
        best_sheet_score = -1

        # Scan all sheets in the workbook and pick the highest scoring schedule sheet
        for sname in wb.sheetnames:
            ws = wb[sname]
            try:
                parsed = _parse_excel_sheet(ws)
                if parsed and parsed.get("tasks"):
                    task_count = len(parsed["tasks"])
                    cols_count = len(parsed.get("columns_detected", {}))
                    dated_count = sum(1 for t in parsed["tasks"] if t.get("start_date") or t.get("due_date"))
                    assigned_count = sum(1 for t in parsed["tasks"] if t.get("assignee_name"))
                    
                    sheet_score = (task_count * 10) + (cols_count * 50) + (dated_count * 30) + (assigned_count * 20)
                    if any(k in sname.lower() for k in ["gantt", "schedule", "production", "plan", "tasks", "activities"]):
                        sheet_score += 100

                    if sheet_score > best_sheet_score:
                        best_sheet_score = sheet_score
                        best_sheet_parsed = parsed
            except Exception:
                continue

        if best_sheet_parsed and best_sheet_parsed.get("tasks"):
            return best_sheet_parsed

        # Fallback to active sheet
        return _parse_excel_sheet(wb.active)
    else:
        return _parse_csv_text(file_bytes)

def _parse_excel_sheet(ws, wb=None):
    rows_raw = []
    for row in ws.iter_rows(values_only=True):
        row_list = list(row)
        if any(c is not None and str(c).strip() != "" for c in row_list):
            rows_raw.append(row_list)

    if not rows_raw:
        raise ValueError("The uploaded Excel sheet contains no data.")

    return _process_table_rows(rows_raw)

def _parse_csv_text(file_bytes):
    text = file_bytes.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    delimiter = "\t" if "\t" in sample and "," not in sample else (";" if ";" in sample and "," not in sample else ",")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows_raw = [row for row in reader if any(c.strip() for c in row)]

    if not rows_raw:
        raise ValueError("The uploaded CSV file is empty.")

    return _process_table_rows(rows_raw)

def _process_table_rows(rows_raw):
    """
    Intelligently extracts tasks, assignees, timelines, and metadata from raw rows.
    Works for any table format (1 column, 2 columns, or full Gantt sheet).
    """
    max_cols = max(len(r) for r in rows_raw) if rows_raw else 0

    # 1. Extract Project Metadata if present in top header banners
    project_metadata = {
        "project_name": None,
        "project_code": None,
        "customer": None,
        "product": None,
        "suggested_title": None
    }
    
    for r in range(min(12, len(rows_raw))):
        row = rows_raw[r]
        for c in range(len(row) - 1):
            k = row[c]
            if k and isinstance(k, str):
                k_clean = k.strip().lower().rstrip(":")
                v = row[c+1]
                if v is not None and str(v).strip():
                    if "project name" in k_clean or k_clean == "project":
                        project_metadata["project_name"] = str(v).strip()
                    elif "project code" in k_clean or k_clean == "code":
                        project_metadata["project_code"] = str(v).strip()
                    elif "customer" in k_clean:
                        project_metadata["customer"] = str(v).strip()
                    elif "product" in k_clean or "target qty" in k_clean:
                        project_metadata["product"] = str(v).strip()

    title_parts = []
    if project_metadata["project_name"]:
        title_parts.append(project_metadata["project_name"])
    if project_metadata["project_code"]:
        title_parts.append(f"({project_metadata['project_code']})")
    if project_metadata["product"]:
        title_parts.append(f"- {project_metadata['product']}")
    
    if title_parts:
        project_metadata["suggested_title"] = " ".join(title_parts)

    # 2. Search for Header Row
    best_header_idx = -1
    best_header_map = {}
    
    for idx in range(min(30, len(rows_raw))):
        row = rows_raw[idx]
        mapping = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            norm = normalize_header(cell)
            field = match_column_name(norm)
            if field and field not in mapping.values():
                mapping[col_idx] = field
        
        # If we match title or any 2 useful columns
        if "title" in mapping.values():
            if len(mapping) > len(best_header_map) or best_header_idx == -1:
                best_header_idx = idx
                best_header_map = mapping
        elif len(mapping) >= 2 and len(mapping) > len(best_header_map):
            best_header_idx = idx
            best_header_map = mapping

    # 3. Fallback: If no recognized header row is found (e.g. raw list of activities)
    if best_header_idx == -1 or not best_header_map or "title" not in best_header_map.values():
        max_cols = max(len(r) for r in rows_raw)
        col_text_scores = [0] * max_cols
        
        for row in rows_raw:
            for c_idx, cell in enumerate(row):
                if cell is not None:
                    c_str = str(cell).strip()
                    if len(c_str) > 2 and not re.match(r"^\d+(\.\d+)?$", c_str) and not parse_date_value(c_str):
                        col_text_scores[c_idx] += 1

        best_title_col = 0
        if col_text_scores and max(col_text_scores) > 0:
            best_title_col = col_text_scores.index(max(col_text_scores))

        row0_val = str(rows_raw[0][best_title_col] or "").strip().lower() if len(rows_raw[0]) > best_title_col else ""
        if any(row0_val == h for h in ["task", "tasks", "title", "activity", "name", "action", "items"]):
            best_header_idx = 0
            data_rows = rows_raw[1:]
        else:
            best_header_idx = -1
            data_rows = rows_raw
        
        best_header_map = {best_title_col: "title"}
        if max_cols > 1:
            for c_idx in range(max_cols):
                if c_idx == best_title_col:
                    continue
                sample_vals = [r[c_idx] for r in data_rows[:10] if c_idx < len(r) and r[c_idx] is not None]
                if any(parse_date_value(v) for v in sample_vals):
                    if "start_date" not in best_header_map.values():
                        best_header_map[c_idx] = "start_date"
                    elif "due_date" not in best_header_map.values():
                        best_header_map[c_idx] = "due_date"
                elif "assignee" not in best_header_map.values() and any(isinstance(v, str) and len(v.strip()) > 1 for v in sample_vals):
                    best_header_map[c_idx] = "assignee"
    else:
        data_rows = rows_raw[best_header_idx + 1:]

    # 4. Detect Weekly Date Ranges in header rows
    week_ranges = {}
    if best_header_idx >= 0:
        date_row_candidates = [best_header_idx - 1, best_header_idx - 2, best_header_idx + 1]
        for d_idx in date_row_candidates:
            if 0 <= d_idx < len(rows_raw):
                row_d = rows_raw[d_idx]
                for col_idx, cell in enumerate(row_d):
                    if cell and isinstance(cell, str) and "to" in cell.lower():
                        parts = re.split(r"\s*to\s*", cell, flags=re.IGNORECASE)
                        if len(parts) == 2:
                            s_d = parse_date_value(parts[0])
                            e_d = parse_date_value(parts[1])
                            if s_d and e_d:
                                week_ranges[col_idx] = (s_d, e_d)

    tasks = []
    members_set = set()
    sprints_set = set()

    title_col = None
    assignee_col = None
    dept_col = None
    status_col = None
    priority_col = None
    desc_col = None
    est_col = None
    sprint_col = None
    start_col = None
    due_col = None
    task_no_col = None

    for col_idx, field in best_header_map.items():
        if field == "title": title_col = col_idx
        elif field == "assignee": assignee_col = col_idx
        elif field == "department": dept_col = col_idx
        elif field == "status": status_col = col_idx
        elif field == "priority": priority_col = col_idx
        elif field == "description": desc_col = col_idx
        elif field == "estimated_hours": est_col = col_idx
        elif field == "sprint": sprint_col = col_idx
        elif field == "start_date": start_col = col_idx
        elif field == "due_date": due_col = col_idx
        elif field == "task_no": task_no_col = col_idx

    current_order_idx = 0
    current_section_name = None

    for row in data_rows:
        if not row or not any(c is not None and str(c).strip() != "" for c in row):
            continue

        raw_title = row[title_col] if title_col is not None and title_col < len(row) else None
        
        # If title column is empty, check if another column has title text
        if not raw_title or str(raw_title).strip() == "":
            for c_idx, cell in enumerate(row):
                if cell is not None and str(cell).strip() != "":
                    if c_idx != task_no_col and not parse_date_value(cell) and not re.match(r"^\d+(\.\d+)?$", str(cell).strip()):
                        raw_title = cell
                        break

        if not raw_title or str(raw_title).strip() == "":
            continue

        title_str = str(raw_title).strip()
        lower_title = title_str.lower()

        # Skip summary, total, or page header lines
        if lower_title in ("total time", "summary", "kpi", "total", "grand total", "sub total", "notes", "legend"):
            continue

        # Dates: Parse explicitly if provided
        start_date = parse_date_value(row[start_col]) if start_col is not None and start_col < len(row) else None
        due_date = parse_date_value(row[due_col]) if due_col is not None and due_col < len(row) else None

        # Assignee
        raw_assignee = row[assignee_col] if assignee_col is not None and assignee_col < len(row) else None
        assignee_str = str(raw_assignee).strip() if raw_assignee is not None else None
        if assignee_str and assignee_str.lower() in ("unassigned", "n/a", "none", "-", "tbd", "undeclared"):
            assignee_str = None

        # Check if this row is a Section / Phase header in a multi-column table (e.g. "Phase 1: Lab Synthesis" or "Sprint 2")
        non_empty_cells = [c for c in row if c is not None and str(c).strip() != ""]
        is_section = False
        if len(non_empty_cells) == 1 and max_cols > 1 and len(best_header_map) >= 2:
            if not start_date and not due_date and not assignee_str:
                if any(k in title_str.lower() for k in ["phase", "sprint", "stage", "section", "module", "part", "milestone", "wbs", "stream"]):
                    is_section = True
                elif max_cols >= 3 and len(title_str) > 2 and not parse_date_value(title_str) and not re.match(r"^\d+$", title_str):
                    is_section = True

        if is_section:
            current_section_name = title_str
            sprints_set.add(current_section_name)
            continue

        if assignee_str:
            sub_members = [m.strip() for m in re.split(r"[,&/]+", assignee_str) if m.strip() and m.strip().lower() not in ("pm team", "scm team", "qa", "qc", "production team")]
            for sm in sub_members:
                members_set.add(sm)
            members_set.add(assignee_str)

        # Department / Role
        raw_dept = row[dept_col] if dept_col is not None and dept_col < len(row) else ""
        dept_str = str(raw_dept).strip() if raw_dept else ""

        # Status & Remarks
        raw_status = row[status_col] if status_col is not None and status_col < len(row) else "todo"
        raw_desc = row[desc_col] if desc_col is not None and desc_col < len(row) else ""
        desc_str = str(raw_desc).strip() if raw_desc else ""

        norm_status = normalize_status(raw_status)
        norm_priority = normalize_priority(
            row[priority_col] if priority_col is not None and priority_col < len(row) else None,
            status=raw_status,
            remark=desc_str
        )

        # Check other date cells or weekly timeline ranges if explicit dates were not found
        if not start_date or not due_date:
            for c_idx in range(len(row)):
                if c_idx in (title_col, assignee_col, dept_col, status_col, desc_col, task_no_col):
                    continue
                cell_val = row[c_idx]
                if cell_val is not None and str(cell_val).strip() != "":
                    if isinstance(cell_val, (datetime, date)):
                        d_str = cell_val.strftime("%Y-%m-%d")
                        if not start_date or d_str < start_date: start_date = d_str
                        if not due_date or d_str > due_date: due_date = d_str
                    elif isinstance(cell_val, str):
                        d_str = parse_date_value(cell_val)
                        if d_str:
                            if not start_date or d_str < start_date: start_date = d_str
                            if not due_date or d_str > due_date: due_date = d_str
                        elif c_idx in week_ranges:
                            s_d, e_d = week_ranges[c_idx]
                            if not start_date or s_d < start_date: start_date = s_d
                            if not due_date or e_d > due_date: due_date = e_d
                    elif c_idx in week_ranges:
                        s_d, e_d = week_ranges[c_idx]
                        if not start_date or s_d < start_date: start_date = s_d
                        if not due_date or e_d > due_date: due_date = e_d

        # Tags
        tags = []
        if dept_str:
            tags.append(dept_str.lower())
        if project_metadata.get("project_code"):
            tags.append(project_metadata["project_code"].lower())

        # Sprint / Phase
        sprint_name = None
        if sprint_col is not None and sprint_col < len(row) and row[sprint_col]:
            sprint_val = str(row[sprint_col]).strip()
            if sprint_val and sprint_val.lower() not in ("none", "null", "-", "n/a", "no sprint", "backlog"):
                sprint_name = sprint_val
                sprints_set.add(sprint_name)
        elif current_section_name:
            sprint_name = current_section_name

        # Estimated Hours
        est_hours = 0.0
        if est_col is not None and est_col < len(row):
            est_hours = normalize_estimated_hours(row[est_col]) or 0.0

        tasks.append({
            "title": title_str,
            "assignee_name": assignee_str,
            "department": dept_str,
            "start_date": start_date,
            "due_date": due_date,
            "status": norm_status,
            "priority": norm_priority,
            "estimated_hours": est_hours,
            "sprint_name": sprint_name,
            "description": desc_str,
            "tags": tags,
            "order_index": current_order_idx
        })
        current_order_idx += 1

    return {
        "tasks": tasks,
        "members_found": sorted(list(members_set)),
        "sprints_found": sorted(list(sprints_set)),
        "total_rows": len(tasks),
        "columns_detected": {v: k for k, v in best_header_map.items()},
        "metadata": project_metadata
    }

def generate_sample_gantt_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Task Name", "Assignee", "Start Date", "Due Date",
        "Status", "Priority", "Estimated Hours", "Sprint / Phase", "Description", "Tags"
    ])
    now = datetime.now()
    d = lambda offset: (now + timedelta(days=offset)).strftime("%Y-%m-%d")
    sample_rows = [
        ["Project Discovery & Scope Alignment", "Alex Morgan", d(-5), d(-2), "Done", "High", "12", "Phase 1: Foundation", "Define architecture and sprint goals", "planning,arch"],
        ["Database Schema & Indices Migration", "Sophia Chen", d(-3), d(3), "In Progress", "Urgent", "16", "Phase 1: Foundation", "SQLite indexing and foreign key audit", "database,backend"],
        ["Interactive Gantt Chart & Kanban UI", "David Kim", d(-1), d(6), "In Progress", "High", "20", "Phase 2: Core Development", "Timeline visualization and drag-and-drop cards", "frontend,ui"],
        ["REST API Authentication Guard", "Alex Morgan", d(2), d(8), "To Do", "Medium", "8", "Phase 2: Core Development", "JWT token and route guard rules", "security,auth"],
        ["Automated End-to-End Test Suite", "Elena Rostova", d(5), d(11), "To Do", "High", "14", "Phase 3: QA & Verification", "Unit test runners and CI verification", "qa,ci"],
        ["Production Staging Deployment", "Marcus Vance", d(9), d(14), "Backlog", "Urgent", "10", "Phase 3: QA & Launch", "Final stage rollout and sanity check", "release,ops"]
    ]
    for r in sample_rows:
        writer.writerow(r)
    return output.getvalue()

def generate_sample_gantt_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt Schedule"
    headers = [
        "Task Name", "Assignee", "Start Date", "Due Date",
        "Status", "Priority", "Estimated Hours", "Sprint / Phase", "Description", "Tags"
    ]
    ws.append(headers)
    now = datetime.now()
    d = lambda offset: (now + timedelta(days=offset)).strftime("%Y-%m-%d")
    sample_rows = [
        ["Project Discovery & Scope Alignment", "Alex Morgan", d(-5), d(-2), "Done", "High", 12, "Phase 1: Foundation", "Define architecture and sprint goals", "planning,arch"],
        ["Database Schema & Indices Migration", "Sophia Chen", d(-3), d(3), "In Progress", "Urgent", 16, "Phase 1: Foundation", "SQLite indexing and foreign key audit", "database,backend"],
        ["Interactive Gantt Chart & Kanban UI", "David Kim", d(-1), d(6), "In Progress", "High", 20, "Phase 2: Core Development", "Timeline visualization and drag-and-drop cards", "frontend,ui"],
        ["REST API Authentication Guard", "Alex Morgan", d(2), d(8), "To Do", "Medium", 8, "Phase 2: Core Development", "JWT token and route guard rules", "security,auth"],
        ["Automated End-to-End Test Suite", "Elena Rostova", d(5), d(11), "To Do", "High", 14, "Phase 3: QA & Verification", "Unit test runners and CI verification", "qa,ci"],
        ["Production Staging Deployment", "Marcus Vance", d(9), d(14), "Backlog", "Urgent", 10, "Phase 3: QA & Launch", "Final stage rollout and sanity check", "release,ops"]
    ]
    for row in sample_rows:
        ws.append(row)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

