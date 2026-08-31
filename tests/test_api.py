import json
import os
import unittest
from datetime import datetime, timezone
import io

TEST_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_pulse.db")
os.environ["PROJECT_PULSE_DB"] = TEST_DB_PATH

from app.database import init_db, get_db
from app.seed import seed_database
from app.main import app
from app.gantt_parser import generate_sample_gantt_csv, generate_sample_gantt_excel

class TestProjectPulseAPI(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        if os.path.exists(TEST_DB_PATH):
            try:
                os.remove(TEST_DB_PATH)
            except Exception:
                pass
        init_db()
        seed_database()

    @classmethod
    def tearDownClass(cls):
        if os.path.exists(TEST_DB_PATH):
            try:
                os.remove(TEST_DB_PATH)
            except Exception:
                pass

    def request(self, path, method="GET", body=None, query=None, raw_body=None, content_type="application/json"):
        if raw_body is not None:
            body_bytes = raw_body
        elif body is not None:
            body_bytes = json.dumps(body).encode("utf-8")
        else:
            body_bytes = b""

        env = {
            "REQUEST_METHOD": method,
            "PATH_INFO": path,
            "SCRIPT_NAME": "",
            "SERVER_NAME": "localhost",
            "SERVER_PORT": "8000",
            "wsgi.version": (1, 0),
            "wsgi.url_scheme": "http",
            "wsgi.input": io.BytesIO(body_bytes),
            "CONTENT_LENGTH": str(len(body_bytes)),
            "CONTENT_TYPE": content_type,
            "QUERY_STRING": query or "",
            "wsgi.errors": io.StringIO(),
            "wsgi.multithread": False,
            "wsgi.multiprocess": False,
            "wsgi.run_once": False,
        }
        
        status_holder = []
        headers_holder = []
        def start_response(status, headers, exc_info=None):
            status_holder.append(status)
            headers_holder.append(headers)

        response_body_chunks = app(env, start_response)
        response_body = b"".join(response_body_chunks)
        
        status_code = int(status_holder[0].split()[0])
        try:
            json_data = json.loads(response_body.decode("utf-8"))
        except Exception:
            json_data = response_body

        return status_code, json_data

    def test_01_get_projects(self):
        status, data = self.request("/api/projects")
        self.assertEqual(status, 200)
        self.assertIsInstance(data, list)
        self.assertGreaterEqual(len(data), 1)
        self.assertIn("CloudScale Platform 2.0", [p["name"] for p in data])

    def test_02_create_and_delete_project(self):
        status, project = self.request("/api/projects", method="POST", body={
            "name": "Automated Test Project",
            "description": "Created during CI test suite",
            "color": "#10B981"
        })
        self.assertEqual(status, 201)
        p_id = project["id"]
        self.assertEqual(project["name"], "Automated Test Project")

        status, members = self.request(f"/api/projects/{p_id}/members")
        self.assertEqual(status, 200)
        self.assertGreaterEqual(len(members), 1)

        status, del_res = self.request(f"/api/projects/{p_id}", method="DELETE")
        self.assertEqual(status, 200)
        self.assertTrue(del_res["success"])

    def test_03_task_lifecycle_and_subtasks(self):
        status, task = self.request("/api/projects/1/tasks", method="POST", body={
            "title": "Build WebSocket Authentication Guard",
            "description": "JWT token verification handshake",
            "status": "todo",
            "priority": "high",
            "estimated_hours": 6.0,
            "tags": ["auth", "security", "websocket"],
            "subtasks": ["Token parser", "Signature validator"]
        })
        self.assertEqual(status, 200)
        t_id = task["id"]
        self.assertEqual(task["title"], "Build WebSocket Authentication Guard")
        self.assertEqual(len(task["subtasks"]), 2)

        sub_id = task["subtasks"][0]["id"]
        status, sub_updated = self.request(f"/api/subtasks/{sub_id}", method="PUT", body={"completed": True})
        self.assertEqual(status, 200)
        self.assertEqual(sub_updated["completed"], 1)

        status, log_res = self.request(f"/api/tasks/{t_id}/timelogs", method="POST", body={
            "hours": 3.5,
            "description": "Implemented token parser logic"
        })
        self.assertEqual(status, 201)
        self.assertEqual(log_res["hours"], 3.5)

        status, task_updated = self.request(f"/api/tasks/{t_id}")
        self.assertEqual(status, 200)
        self.assertEqual(task_updated["actual_hours"], 3.5)

        self.request(f"/api/tasks/{t_id}", method="DELETE")

    def test_04_drag_and_drop_reorder(self):
        status, res = self.request("/api/tasks/reorder", method="POST", body=[
            {"task_id": 1, "status": "in_progress", "new_order_index": 0},
            {"task_id": 2, "status": "done", "new_order_index": 1}
        ])
        self.assertEqual(status, 200)
        self.assertTrue(res["success"])

    def test_05_analytics_and_burndown(self):
        status, analytics = self.request("/api/projects/1/analytics")
        self.assertEqual(status, 200)
        self.assertIn("kpis", analytics)
        self.assertIn("burndown", analytics)
        self.assertIn("priority_distribution", analytics)
        self.assertIn("workload", analytics)
        self.assertGreater(analytics["kpis"]["total_tasks"], 0)

    def test_06_project_export_and_import(self):
        status, export_data = self.request("/api/projects/1/export")
        self.assertEqual(status, 200)
        self.assertEqual(export_data["version"], "1.0")
        self.assertIn("tasks", export_data)

        status, import_res = self.request("/api/projects/import", method="POST", body=export_data)
        self.assertEqual(status, 200)
        self.assertTrue(import_res["success"])
        new_p_id = import_res["project_id"]

        self.request(f"/api/projects/{new_p_id}", method="DELETE")

    def test_07_gantt_sample_downloads(self):
        status, csv_data = self.request("/api/gantt/sample_csv")
        self.assertEqual(status, 200)
        self.assertIn(b"Task Name", csv_data)

        status, xlsx_data = self.request("/api/gantt/sample_xlsx")
        self.assertEqual(status, 200)
        self.assertTrue(xlsx_data.startswith(b"PK\x03\x04"))

    def test_08_upload_gantt_multipart(self):
        # Build multipart/form-data payload with Excel file
        xlsx_bytes = generate_sample_gantt_excel()
        boundary = "----WebKitFormBoundary7MA4YWxkTrZu0gW"
        
        body_parts = []
        body_parts.append(f"--{boundary}\r\n".encode("utf-8"))
        body_parts.append(f'Content-Disposition: form-data; name="file"; filename="gantt_schedule.xlsx"\r\n'.encode("utf-8"))
        body_parts.append(b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n")
        body_parts.append(xlsx_bytes)
        body_parts.append(b"\r\n")
        body_parts.append(f"--{boundary}\r\n".encode("utf-8"))
        body_parts.append(f'Content-Disposition: form-data; name="new_project_name"\r\n\r\n'.encode("utf-8"))
        body_parts.append(b"Uploaded Gantt Milestone Project\r\n")
        body_parts.append(f"--{boundary}--\r\n".encode("utf-8"))
        
        multipart_body = b"".join(body_parts)
        
        status, res = self.request(
            "/api/projects/1/upload_gantt",
            method="POST",
            raw_body=multipart_body,
            content_type=f"multipart/form-data; boundary={boundary}"
        )
        self.assertEqual(status, 200)
        self.assertTrue(res["success"])
        self.assertGreaterEqual(res["tasks_imported"], 5)
        self.assertGreaterEqual(res["members_added"], 1)

        # Cleanup imported project
        new_p_id = res["project_id"]
        self.request(f"/api/projects/{new_p_id}", method="DELETE")

    def test_09_edit_all_fields(self):
        # Create a new task
        status, task = self.request("/api/projects/1/tasks", method="POST", body={
            "title": "Initial Task Title",
            "description": "Initial description",
            "status": "todo",
            "priority": "low",
            "start_date": "2026-08-10",
            "due_date": "2026-08-15",
            "estimated_hours": 4.0
        })
        self.assertEqual(status, 200)
        t_id = task["id"]

        # 1. Update Start and End / Due Dates
        status, updated = self.request(f"/api/tasks/{t_id}", method="PUT", body={
            "start_date": "2026-08-12",
            "due_date": "2026-08-20"
        })
        self.assertEqual(status, 200)
        self.assertEqual(updated["start_date"], "2026-08-12")
        self.assertEqual(updated["due_date"], "2026-08-20")

        # 2. Inline update title, priority, status, hours
        status, updated = self.request(f"/api/tasks/{t_id}", method="PUT", body={
            "title": "Updated Task Title with All Editable Fields",
            "priority": "urgent",
            "status": "in_progress",
            "estimated_hours": 12.5,
            "actual_hours": 8.0
        })
        self.assertEqual(status, 200)
        self.assertEqual(updated["title"], "Updated Task Title with All Editable Fields")
        self.assertEqual(updated["priority"], "urgent")
        self.assertEqual(updated["status"], "in_progress")
        self.assertEqual(updated["estimated_hours"], 12.5)
        self.assertEqual(updated["actual_hours"], 8.0)

        # 3. Test clearing dates (setting to empty string or None)
        status, cleared = self.request(f"/api/tasks/{t_id}", method="PUT", body={
            "start_date": "",
            "due_date": None
        })
        self.assertEqual(status, 200)
        self.assertIsNone(cleared["start_date"])
        self.assertIsNone(cleared["due_date"])

        # Clean up
        self.request(f"/api/tasks/{t_id}", method="DELETE")

    def test_10_email_settings_and_test_email(self):
        # 1. Get default email settings
        status, settings = self.request("/api/notifications/settings")
        self.assertEqual(status, 200)
        self.assertIn("smtp_host", settings)
        self.assertEqual(settings["is_enabled"], 1)

        # 2. Update email settings
        status, updated = self.request("/api/notifications/settings", method="PUT", body={
            "smtp_provider": "outlook",
            "smtp_host": "smtp.office365.com",
            "smtp_port": 587,
            "smtp_user": "tester@outlook.com",
            "smtp_pass": "secretAppPass123",
            "sender_name": "Test Workspace",
            "simulation_mode": 1,
            "notify_due_soon": 1,
            "notify_due_today": 1,
            "notify_overdue": 1,
            "recurring_day": "monday"
        })
        self.assertEqual(status, 200)
        self.assertEqual(updated["smtp_user"], "tester@outlook.com")
        self.assertEqual(updated["sender_name"], "Test Workspace")
        self.assertTrue(updated["has_password"])

        # 3. Send test email in simulation mode
        status, test_res = self.request("/api/notifications/test_email", method="POST", body={
            "email": "recipient@company.internal"
        })
        self.assertEqual(status, 200)
        self.assertTrue(test_res["success"])
        self.assertEqual(test_res["status"], "simulated")

        # 4. Check outbox logs
        status, logs = self.request("/api/notifications/logs")
        self.assertEqual(status, 200)
        self.assertGreaterEqual(len(logs), 1)
        self.assertEqual(logs[0]["recipient_email"], "recipient@company.internal")
        self.assertEqual(logs[0]["trigger_type"], "test_email")

    def test_11_assignment_and_completion_triggers(self):
        # 1. Create a test member with email
        status, member = self.request("/api/projects/1/members", method="POST", body={
            "name": "Jordan Lee",
            "email": "jordan.lee@example.com",
            "role": "QA Engineer",
            "avatar_color": "#10B981"
        })
        self.assertIn(status, [200, 201])
        m_id = member["id"]

        # 2. Create task assigned to member -> Action 4 (Task Assigned)
        status, task = self.request("/api/projects/1/tasks", method="POST", body={
            "title": "Azadol Spectroscopic Purity Analysis",
            "status": "todo",
            "priority": "high",
            "assignee_id": m_id,
            "start_date": "2026-08-28",
            "due_date": "2026-09-05"
        })
        self.assertIn(status, [200, 201])
        t_id = task["id"]

        # Verify assignment email logged
        status, logs = self.request("/api/notifications/logs")
        self.assertEqual(status, 200)
        assign_logs = [l for l in logs if l["task_id"] == t_id and l["trigger_type"] == "assigned"]
        self.assertGreaterEqual(len(assign_logs), 1)
        self.assertEqual(assign_logs[0]["recipient_email"], "jordan.lee@example.com")

        # 3. Update task to 'done' -> Action 5 (Task Completed)
        status, updated_task = self.request(f"/api/tasks/{t_id}", method="PUT", body={
            "status": "done"
        })
        self.assertEqual(status, 200)

        # Verify completion email logged
        status, logs = self.request("/api/notifications/logs")
        complete_logs = [l for l in logs if l["task_id"] == t_id and l["trigger_type"] == "completed"]
        self.assertGreaterEqual(len(complete_logs), 1)

        # Clean up
        self.request(f"/api/tasks/{t_id}", method="DELETE")

    def test_12_due_date_checks_and_deduplication(self):
        # 1. Create a member with email
        status, member = self.request("/api/projects/1/members", method="POST", body={
            "name": "Devin Taskmaster",
            "email": "devin.master@example.com",
            "role": "Lead",
            "avatar_color": "#F59E0B"
        })
        self.assertIn(status, [200, 201])
        m_id = member["id"]

        # 2. Create tasks:
        # Task A: Due tomorrow (2026-08-29) -> Action 1
        # Task B: Due today (2026-08-28) -> Action 2
        # Task C: Due yesterday (2026-08-27) -> Action 3 (1 day overdue)
        status, task_a = self.request("/api/projects/1/tasks", method="POST", body={
            "title": "Task Due Tomorrow",
            "status": "in_progress",
            "assignee_id": m_id,
            "due_date": "2026-08-29"
        })
        status, task_b = self.request("/api/projects/1/tasks", method="POST", body={
            "title": "Task Due Today",
            "status": "todo",
            "assignee_id": m_id,
            "due_date": "2026-08-28"
        })
        status, task_c = self.request("/api/projects/1/tasks", method="POST", body={
            "title": "Task Overdue 1 Day",
            "status": "todo",
            "assignee_id": m_id,
            "due_date": "2026-08-27"
        })

        # 3. Trigger check for reference date 2026-08-28
        status, res = self.request("/api/notifications/run_checks", method="POST", body={
            "reference_date": "2026-08-28"
        })
        self.assertEqual(status, 200)
        self.assertGreaterEqual(res["notifications_dispatched"], 3)

        # 4. Re-run check on same date -> Deduplication should dispatch 0 new notifications
        status, res2 = self.request("/api/notifications/run_checks", method="POST", body={
            "reference_date": "2026-08-28"
        })
        self.assertEqual(status, 200)
        # Deduplication should ensure no duplicates on the same date for these tasks
        self.assertEqual(res2["notifications_dispatched"], 0)

        # Clean up
        self.request(f"/api/tasks/{task_a['id']}", method="DELETE")
        self.request(f"/api/tasks/{task_b['id']}", method="DELETE")
        self.request(f"/api/tasks/{task_c['id']}", method="DELETE")

    def test_13_universal_spreadsheet_import(self):
        from app.gantt_parser import parse_gantt_file

        # Test 1: Basic 1-column list of activity names
        csv_1col = b"Reaction Step 1\nReaction Step 2\nReaction Step 3\n"
        res1 = parse_gantt_file(csv_1col, "activities.csv")
        self.assertEqual(len(res1["tasks"]), 3)
        self.assertEqual(res1["tasks"][0]["title"], "Reaction Step 1")
        self.assertIsNone(res1["tasks"][0]["start_date"])
        self.assertIsNone(res1["tasks"][0]["due_date"])

        # Test 2: Basic 2-column list (Activity, Owner)
        csv_2col = b"Activity,Owner\nRaw Material Charging,Sri Vishnu\nYield Check,QA Lead\n"
        res2 = parse_gantt_file(csv_2col, "steps.csv")
        self.assertEqual(len(res2["tasks"]), 2)
        self.assertEqual(res2["tasks"][0]["title"], "Raw Material Charging")
        self.assertEqual(res2["tasks"][0]["assignee_name"], "Sri Vishnu")
        self.assertIsNone(res2["tasks"][0]["start_date"])

        # Test 3: Excel with custom headers, named month dates, and serial dates
        import openpyxl, io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Deliverables", "Status", "Responsible Person", "Planned Start", "Target Finish"])
        ws.append(["Phase 1: Foundation", None, None, None, None]) # Section header
        ws.append(["API Architecture Spec", "Done", "Elena", "15-Aug-2026", "25-Aug-2026"])
        ws.append(["Database Migration", "In Progress", "Alex", 45520, 45530]) # Serial numbers
        buf = io.BytesIO()
        wb.save(buf)

        res3 = parse_gantt_file(buf.getvalue(), "project_tasks.xlsx")
        self.assertEqual(len(res3["tasks"]), 2)
        self.assertEqual(res3["tasks"][0]["title"], "API Architecture Spec")
        self.assertEqual(res3["tasks"][0]["status"], "done")
        self.assertEqual(res3["tasks"][0]["assignee_name"], "Elena")
        self.assertEqual(res3["tasks"][0]["start_date"], "2026-08-15")
        self.assertEqual(res3["tasks"][0]["due_date"], "2026-08-25")
        self.assertEqual(res3["tasks"][0]["sprint_name"], "Phase 1: Foundation")

        # Serial dates check
        self.assertEqual(res3["tasks"][1]["title"], "Database Migration")
        self.assertEqual(res3["tasks"][1]["start_date"], "2024-08-16")
        self.assertEqual(res3["tasks"][1]["due_date"], "2024-08-26")

    def test_14_authentication_flow(self):
        # 1. Invalid credentials -> 401
        status, res = self.request("/api/auth/login", method="POST", body={
            "username": "admin",
            "password": "wrongpassword"
        })
        self.assertEqual(status, 401)

        # 2. Valid Login as admin -> 200
        status, login_res = self.request("/api/auth/login", method="POST", body={
            "username": "admin",
            "password": "admin123"
        })
        self.assertEqual(status, 200)
        self.assertTrue(login_res["success"])
        token = login_res["token"]
        self.assertEqual(login_res["user"]["username"], "admin")
        self.assertEqual(login_res["user"]["role"], "admin")

        # 3. GET /api/auth/me with Bearer token -> 200
        env = {
            "REQUEST_METHOD": "GET",
            "PATH_INFO": "/api/auth/me",
            "SCRIPT_NAME": "",
            "SERVER_NAME": "localhost",
            "SERVER_PORT": "8000",
            "wsgi.version": (1, 0),
            "wsgi.url_scheme": "http",
            "wsgi.input": io.BytesIO(b""),
            "CONTENT_LENGTH": "0",
            "CONTENT_TYPE": "application/json",
            "QUERY_STRING": "",
            "HTTP_AUTHORIZATION": f"Bearer {token}",
            "wsgi.errors": io.StringIO(),
            "wsgi.multithread": False,
            "wsgi.multiprocess": False,
            "wsgi.run_once": False,
        }
        status_holder = []
        headers_holder = []
        def start_resp(st, hd, exc=None):
            status_holder.append(st)
            headers_holder.append(hd)
        body_chunks = app(env, start_resp)
        me_res = json.loads(b"".join(body_chunks).decode("utf-8"))
        self.assertEqual(int(status_holder[0].split()[0]), 200)
        self.assertTrue(me_res["authenticated"])
        self.assertEqual(me_res["user"]["username"], "admin")

        # 4. Register new user -> 201
        status, reg_res = self.request("/api/auth/register", method="POST", body={
            "full_name": "Test User",
            "username": "testuser_auth",
            "email": "testuser@company.internal",
            "password": "secretPassword123"
        })
        self.assertEqual(status, 201)
        self.assertTrue(reg_res["success"])
        self.assertEqual(reg_res["user"]["username"], "testuser_auth")

        # 5. Duplicate registration -> 400
        status, dup_res = self.request("/api/auth/register", method="POST", body={
            "full_name": "Test User 2",
            "username": "testuser_auth",
            "email": "different@company.internal",
            "password": "secretPassword123"
        })
        self.assertEqual(status, 400)
        self.assertIn("already registered", dup_res["error"])

        # 6. Logout
        env_logout = {
            "REQUEST_METHOD": "POST",
            "PATH_INFO": "/api/auth/logout",
            "SCRIPT_NAME": "",
            "SERVER_NAME": "localhost",
            "SERVER_PORT": "8000",
            "wsgi.version": (1, 0),
            "wsgi.url_scheme": "http",
            "wsgi.input": io.BytesIO(b"{}"),
            "CONTENT_LENGTH": "2",
            "CONTENT_TYPE": "application/json",
            "QUERY_STRING": "",
            "HTTP_AUTHORIZATION": f"Bearer {token}",
            "wsgi.errors": io.StringIO(),
            "wsgi.multithread": False,
            "wsgi.multiprocess": False,
            "wsgi.run_once": False,
        }
        status_holder = []
        def start_resp_lo(st, hd, exc=None): status_holder.append(st)
        app(env_logout, start_resp_lo)
        self.assertEqual(int(status_holder[0].split()[0]), 200)

    def test_15_multi_sheet_excel_parsing(self):
        import openpyxl, io
        from app.gantt_parser import parse_gantt_file
        wb = openpyxl.Workbook()
        
        # Sheet 1: Empty instructions
        ws_info = wb.active
        ws_info.title = "Instructions"
        ws_info.append(["Project Management Guidelines"])
        ws_info.append(["Follow ISO standards for all deliverables"])

        # Sheet 2: The actual Gantt schedule
        ws_gantt = wb.create_sheet(title="Production Schedule")
        ws_gantt.append(["Process Step", "Lead", "Target Start", "Target Finish", "Condition"])
        ws_gantt.append(["Chemical Formulation", "Sri Vishnu", "12-Sep-2026", "18-Sep-2026", "In Progress"])
        ws_gantt.append(["HPLC Purity Analysis", "Elena", "20-Sep-2026", "24-Sep-2026", "To Do"])
        
        buf = io.BytesIO()
        wb.save(buf)

        res = parse_gantt_file(buf.getvalue(), "multi_sheet_project.xlsx")
        self.assertEqual(len(res["tasks"]), 2)
        self.assertEqual(res["tasks"][0]["title"], "Chemical Formulation")
        self.assertEqual(res["tasks"][0]["start_date"], "2026-09-12")
        self.assertEqual(res["tasks"][0]["due_date"], "2026-09-18")
        self.assertEqual(res["tasks"][0]["status"], "in_progress")
        self.assertEqual(res["tasks"][1]["title"], "HPLC Purity Analysis")
        self.assertEqual(res["tasks"][1]["assignee_name"], "Elena")

    def test_16_chemical_batch_gantt_schedule_parsing(self):
        import openpyxl, io
        from app.gantt_parser import parse_gantt_file
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Production Schedule"
        ws.append(["Project Name:", "AZADOL (CCS079)", "Customer:", "Chemtatva", "Target Qty:", "1kg Batch Production"])
        ws.append([])
        ws.append(["Task", "Activity / Process Step", "Responsible", "Task Owner", "Start Date", "End Date", "Status", "Remarks", "01/08 to 07/08", "08/08 to 14/08", "15/08 to 21/08"])
        ws.append([1, "PO Received date", "SCM", "Sri Vishnu", "06/08/2026", "06/08/2026", "Done", "PO #441", 8, None, None])
        ws.append([2, "Kickoff Meeting", "PM", "Sri Vishnu", "12/08/2026", "12/08/2026", "Done", "Kickoff complete", None, 8, None])
        ws.append([3, "Raw Material Delivery", "SCM", "Sri Vishnu", "28/08/2026", "28/08/2026", "In Progress", "En route", None, None, 16])
        
        buf = io.BytesIO()
        wb.save(buf)

        res = parse_gantt_file(buf.getvalue(), "batch_schedule.xlsx")
        self.assertEqual(len(res["tasks"]), 3)
        self.assertEqual(res["tasks"][0]["title"], "PO Received date")
        self.assertEqual(res["tasks"][0]["assignee_name"], "Sri Vishnu")
        self.assertEqual(res["tasks"][0]["department"], "SCM")
        self.assertEqual(res["tasks"][0]["start_date"], "2026-08-06")
        self.assertEqual(res["tasks"][0]["due_date"], "2026-08-06")
        self.assertEqual(res["tasks"][0]["status"], "done")

        self.assertEqual(res["tasks"][1]["title"], "Kickoff Meeting")
        self.assertEqual(res["tasks"][1]["status"], "done")
        self.assertEqual(res["tasks"][2]["title"], "Raw Material Delivery")
        self.assertEqual(res["tasks"][2]["status"], "in_progress")

if __name__ == "__main__":
    unittest.main()


